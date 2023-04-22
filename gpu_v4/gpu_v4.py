import cnocr
import re
from PIL import Image
import openpyxl
import os
import datetime
import os
import threading
import queue
import time

def get_buy_volume(buy_volume_crop_img,ocr):
    gray_img = buy_volume_crop_img.convert('L')

    buy_volume_results = []
    text = ocr.ocr(gray_img)
    flag = 0
    """ 要求text数量为8(4对数据)否则代表数据不完整全都赋予-999  """
    if len(text) == 8:
        for i in text:
            flag +=1
            my_str = i["text"]
            """ 300万以上  0.00  后者是目标，所以只取偶数值 """
            if flag % 2 == 0:   
                # if my_str.endswith("万") or my_str.endswith("亿"):
                #     pattern = r'^([-+]?\d+\.\d+|\d+)'  # 匹配数字和符号部分的正则表达式
                #     match = re.match(pattern, my_str)
                if "万" in my_str or "亿" in my_str or "斤" in my_str:
                    pattern = r'([-+]?\d+\.\d+|\d+)(万|亿|斤)'  # 匹配数字和"万"或"亿"部分的正则表达式
                    match = re.match(pattern, my_str)
                    if match:
                        num = match.group(1)
                        buy_volume_results.append(float(num))
                    else:
                        buy_volume_results.append(-999)
                else:
                    buy_volume_results.append(float(my_str))
    else:
        buy_volume_results.append(-999)
        buy_volume_results.append(-999)
        buy_volume_results.append(-999)
        buy_volume_results.append(-999)
    return buy_volume_results


def get_code(code_crop_img,ocr):
    # 获取图像大小
    width, height = code_crop_img.size
    # 查找背景颜色是(0, 0, 128)的像素
    bg_mask = Image.new(mode="1", size=(width, height))
    for x in range(width):
        for y in range(height):
            pixel = code_crop_img.getpixel((x, y))
            if pixel == (0, 0, 128):
                bg_mask.putpixel((x, y), 1)

    # 获取包含背景像素的最小矩形区域
    bbox = bg_mask.getbbox()
    x1 = bbox[0] - 100
    y1 = bbox[1] 
    x2 = bbox[2] 
    y2 = bbox[3] 
    target_box = (x1,y1,x2,y2)
    # 显示包含背景像素的最小矩形区域
    target_code= code_crop_img.crop(target_box)
    # target_code.show()
    # 将裁剪后的图片转换为灰度图像
    code_gray = target_code.convert('L')
    text = ocr.ocr(code_gray)
    code = -999
    if len(text) == 2: # [代码 , 名称 ]
        print(text)
        code = int(text[0]["text"])
    return code 

def get_deal(deal_crop_img,ocr):
    # 将裁剪后的图片转换为灰度图像
    gray_img = deal_crop_img.convert('L')

    deal_results = []
    text = ocr.ocr(gray_img)
    for i in text:
        my_str = i["text"]
        if my_str.endswith("笔"):
            numberslist = re.findall(r'\d+', my_str)
            if len(numberslist) == 0:
                numberslist = ['-999']
            numbers = float(''.join(numberslist))
            deal_results.append(numbers)
    return deal_results

def get_other(filename,formatted_time):
    index = filename.index("-")  # 找到"-"的索引位置
    num = float(filename[:index])  # 获取"-"之前的数字部分
    time_str = filename[index+1:]  # 获取"-"之后的时间部分

    # 将时间字符串转换为"8:22:18"的格式
    time_parts = time_str.split("点")
    hour = time_parts[0]
    time_parts = time_parts[1].split("分")
    minute = time_parts[0]
    time_parts = time_parts[1].split("秒")
    second = time_parts[0]
    time_str = "{}:{}:{}".format(hour, minute, second)
    date_str = formatted_time.replace("-", "/")

    return num,time_str,date_str


# 定义一个函数来处理单个图片
def process_image(image_path, ocr,formatted_time):
    # ... 图片处理的代码 ...
    # 获取文件名和扩展名
    file_name = os.path.basename(image_path)
    try:
        NUM , TIME , DATE= get_other(file_name,formatted_time)
        img = Image.open(image_path)
        # 裁剪图片，仅保留需要识别的区域
        code_crop_img = img.crop((1180, 110, 1430, 965))
        code=get_code(code_crop_img,ocr)
        # print(code)

        # 裁剪图片，仅保留需要识别的区域
        deal_crop_img = img.crop((590, 140, 790, 340))
        deal_results=get_deal(deal_crop_img,ocr)
        # print(deal_results)

        buy_volume_crop_img = img.crop((592, 370, 785, 550))
        buy_volume_results = get_buy_volume(buy_volume_crop_img,ocr)
    except:
        raise Exception(image_path,"识别异常，请检查该图片")
    # print(buy_volume_results)
    print(image_path,"completed")
    result={
        "NUM":NUM,
        "TIME":TIME,
        "DATE":DATE,
        "code":code,
        "deal_results":deal_results,
        "buy_volume_results":buy_volume_results
    }
    return result

# 定义一个线程类来处理图片
class ImageProcessor(threading.Thread):
    def __init__(self, image_paths,result_queue,formatted_time):
        super().__init__()
        self.image_paths = image_paths
        self.ocr = cnocr.CnOcr(context="gpu")
        self.result_queue = result_queue
        self.formatted_time = formatted_time

    def run(self):
        for image_path in self.image_paths:
            result = process_image(image_path, self.ocr,self.formatted_time)
            self.result_queue.put(result)

def new_excel():
    # 创建一个新的 Excel 文件
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 读取指定文件夹下的所有图片，并依次处理并写入 Excel 文件中
    # 插入表头行
    row = 1
    sheet.insert_rows(row)

    # 设置表头内容
    sheet.cell(row=1, column=1).value = 'CODE'
    sheet.cell(row=1, column=2).value = 'B1'
    sheet.cell(row=1, column=3).value = 'S1'
    sheet.cell(row=1, column=4).value = 'B2'
    sheet.cell(row=1, column=5).value = 'S2'
    sheet.cell(row=1, column=6).value = 'B3'
    sheet.cell(row=1, column=7).value = 'S3'
    sheet.cell(row=1, column=8).value = 'B4'
    sheet.cell(row=1, column=9).value = 'S4'
    sheet.cell(row=1, column=10).value = 'UP300'
    sheet.cell(row=1, column=11).value = 'UP100-300'
    sheet.cell(row=1, column=12).value = 'UP50-100'
    sheet.cell(row=1, column=13).value = 'UP30-50'
    sheet.cell(row=1, column=14).value = 'TIME'
    sheet.cell(row=1, column=15).value = 'DATE'
    sheet.cell(row=1, column=16).value = 'NUM'
    return sheet,workbook

def write_excel(sheet,result,row):
    # 将图片名称和识别结果写入 Excel 文件中
    code = result["code"]
    deal_results = result["deal_results"]
    buy_volume_results = result["buy_volume_results"]
    TIME = result["TIME"]
    DATE = result["DATE"]
    NUM = result["NUM"]
    sheet.cell(row=row, column=1, value=code)

    col = 2  # 从第二列开始写入数据
    for deal in deal_results:
        sheet.cell(row=row, column=col, value=deal)
        col += 1
    for buy_volume in buy_volume_results:
        sheet.cell(row=row, column=col, value=buy_volume)
        col += 1
    #TIME
    sheet.cell(row=row, column=col, value=TIME)
    col += 1
    #DATE
    sheet.cell(row=row, column=col, value=DATE)
    col += 1
    #NUM
    sheet.cell(row=row, column=col, value=NUM)


def save_excel(workbook,formatted_time):
    # 保存 Excel 文件
    f_name = formatted_time.replace("-", "")
    folder_path = "翻译"+ f_name
    file_name = "文本" + formatted_time + ".xlsx"
    file_path = os.path.join(folder_path, file_name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    workbook.save(file_path)


def convert(folder_path,formatted_time):
    sheet,workbook = new_excel()
    # 初始化cnocr模型
    # ocr = cnocr.CnOcr()
    image_paths = []
    for filename in os.listdir(folder_path):
        if filename.endswith(".bmp"):
            image_paths.append(os.path.join(folder_path, filename))
    # ... 获取图片列表和初始化 OCR 对象 ...
    # 创建一个队列来存储处理结果
    result_queue = queue.Queue()
     # 将图片列表按照数量平均分成 2 个子列表
    image_lists = [image_paths[i:i+len(image_paths)//2] for i in range(0, len(image_paths), len(image_paths)//2)]
     # 创建 2 个线程来处理图片
    threads = [ImageProcessor(image_list,result_queue,formatted_time) for image_list in image_lists]
    # 启动线程并等待完成
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()

     # 从队列中获取处理结果并写入 Excel 文件
    row = 1
    for i in range(result_queue.qsize()):
        result = result_queue.get()
        # ... 将结果写入 Excel 文件 ...
        row +=1
        write_excel(sheet,result,row)
    # 保存excel
    save_excel(workbook,formatted_time)


 
def get_file(choose_type,formatted_time):
    path = os.getcwd()  # 获取当前运行脚本所在的目录
    dir_list = os.listdir(path)  # 获取目录下所有文件和文件夹的名称列表
    dir_list = [os.path.join(path, d) for d in dir_list if os.path.isdir(os.path.join(path, d)) and d.startswith("截图")] # 输出以"截图"开头的文件夹路径列表
    # 获取文件夹名称列表
    name_list = [os.path.basename(d) for d in dir_list]
    name = "截图"+formatted_time
    if choose_type == "1":
        if name in name_list:# 从name_list 找 “截图”+formatted_time
            return name
        else:
            raise Exception("没有该文件夹:",name)
    elif choose_type == "2":
        return name_list
    elif choose_type == "3":
        if name in name_list:# 从name_list 找 “截图”+formatted_time
            return name
        else:
            raise Exception("没有该文件夹:",name)


def main():
    choose_type = input("请输入类型:1当日数据,2全部数据,3选择日期,按下回车默认使用当日数据):").strip()
    if choose_type == "" or choose_type=="1":
        now = datetime.datetime.now() # 获取当前系统时间
        formatted_time = now.strftime("%Y-%m-%d") # 将时间格式化成字符串
        folder_path=get_file("1",formatted_time)
        convert(folder_path,formatted_time)
        print("success")
    elif choose_type == "2" :
        formatted_time=""
        folder_paths = get_file("2",formatted_time)
        date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')  # 定义日期字符串的正则表达式
        for folder_path in folder_paths:
            formatted_time=folder_path[2:]
            if date_pattern.match(formatted_time):
                convert(folder_path,formatted_time)
        print("success")
    elif choose_type == "3" :
        date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')  # 定义日期字符串的正则表达式
        formatted_time = input("请输入日期(格式为 yyyy-mm-dd):").strip()
        if formatted_time != "" and date_pattern.match(formatted_time) :
            folder_path = get_file("3",formatted_time)
            convert(folder_path,formatted_time)
            print("success")
        else:
            raise Exception("时间格式错误")
    else:
        raise Exception("请检查输入类型")

if __name__ == '__main__':
    start_time = time.time()
    main()
    end_time =time.time()
    print("start_time",start_time)
    print("end_time",end_time)
    print("run_time",end_time-start_time)
