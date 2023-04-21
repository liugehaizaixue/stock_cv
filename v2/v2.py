import cnocr
import re
from PIL import Image
import openpyxl
import os
import warnings
import datetime
import csv


def get_code(code_crop_img,ocr):
    # 获取图像大小
    width, height = code_crop_img.size
    # 查找背景颜色是(0, 0, 128)的像素
    bg_mask = Image.new(mode="1", size=(width, height))
    for x in range(width):
        for y in range(height):
            pixel = code_crop_img.getpixel((x, y))
            if pixel == (0, 0, 128):
                bg_mask.putpixel((x, y), 1)

    # 获取包含背景像素的最小矩形区域
    bbox = bg_mask.getbbox()
    x1 = bbox[0] - 100
    y1 = bbox[1] 
    x2 = bbox[2] 
    y2 = bbox[3] 
    target_box = (x1,y1,x2,y2)
    # 显示包含背景像素的最小矩形区域
    target_code= code_crop_img.crop(target_box)
    # target_code.show()
    # 将裁剪后的图片转换为灰度图像
    code_gray = target_code.convert('L')
    text = ocr.ocr(code_gray)
    code = -999
    if len(text) == 2: # [代码 , 名称 ]
        code = int(text[0]["text"])
    return code 

def convert(formatted_time):

  warnings.filterwarnings('ignore')
  # 指定需要读取的文件夹路径
  folder_path = "截图"+formatted_time

  # 初始化cnocr模型
  ocr = cnocr.CnOcr()

  # 创建一个新的 Excel 文件
  workbook = openpyxl.Workbook()
  sheet = workbook.active

  # 读取指定文件夹下的所有图片，并依次处理并写入 Excel 文件中
  # 插入表头行
  row = 1
  sheet.insert_rows(row)

  # 设置表头内容
  sheet.cell(row=1, column=1).value = '代码'
  sheet.cell(row=1, column=2).value = '名称'
  sheet.cell(row=1, column=3).value = '涨幅%'
  sheet.cell(row=1, column=4).value = '现价'
  sheet.cell(row=1, column=5).value = '涨速%'
  sheet.cell(row=1, column=6).value = '未匹配量'
  sheet.cell(row=1, column=7).value = '今开'
  sheet.cell(row=1, column=8).value = '时长'
  sheet.cell(row=1, column=9).value = '硬盘比'
  sheet.cell(row=1, column=10).value = '红绿比'
  sheet.cell(row=1, column=11).value = 'BS1'
  sheet.cell(row=1, column=12).value = 'BS2'
  sheet.cell(row=1, column=13).value = 'BS3'
  sheet.cell(row=1, column=14).value = 'BS4'

  sheet.cell(row=1, column=16).value = 'B'
  sheet.cell(row=1, column=17).value = 'C'
  sheet.cell(row=1, column=18).value = 'D'
  sheet.cell(row=1, column=19).value = 'E'
  sheet.cell(row=1, column=20).value = 'F'
  sheet.cell(row=1, column=21).value = 'G'
  sheet.cell(row=1, column=22).value = 'H'
  sheet.cell(row=1, column=23).value = 'I'

  for filename in os.listdir(folder_path):
      row = int(filename[:-4])+1
      if filename.endswith(".bmp"):
          # 读取原始图片
          img = Image.open(os.path.join(folder_path, filename))
        
          code_crop_img = img.crop((1180, 110, 1430, 965))
          code=get_code(code_crop_img,ocr)
          # 裁剪图片，仅保留需要识别的区域
          deal_crop_img = img.crop((590, 140, 790, 340))

          # 将裁剪后的图片转换为灰度图像
          gray_img = deal_crop_img.convert('L')

          # 调用cnocr识别文字
          results = []
          text = ocr.ocr(gray_img)
          for i in text:
              my_str = i["text"]
              if my_str.endswith("笔"):
                numberslist = re.findall(r'\d+', my_str)
                if len(numberslist) == 0:
                    numberslist = ['-999']
                numbers = float(''.join(numberslist))
                results.append(numbers)

          # 将图片名称和识别结果写入 Excel 文件中
          col = 16  # 从第二列开始写入数据
          for result in results:
              sheet.cell(row=row, column=1, value=code)
              sheet.cell(row=row, column=col, value=result)
              # 将单元格格式设置为数字格式
              col += 1
          sheet.cell(row=row, column=11, value=results[0]-results[1])
          sheet.cell(row=row, column=12, value=results[2]-results[3])
          sheet.cell(row=row, column=13, value=results[4]-results[5])
          sheet.cell(row=row, column=14, value=results[6]-results[7])


  # 保存 Excel 文件
  folder_path = "翻译"+formatted_time 
  file_name = "创业板" + formatted_time + ".xlsx"
  file_path = os.path.join(folder_path, file_name)
  if not os.path.exists(folder_path):
      os.makedirs(folder_path)
  workbook.save(file_path)




def main():
    date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')  # 定义日期字符串的正则表达式
    formatted_time = input("请输入日期(格式为 yyyy-mm-dd,按下回车默认使用当前日期):").strip()
    if formatted_time == "":
        now = datetime.datetime.now() # 获取当前系统时间
        formatted_time = now.strftime("%Y-%m-%d") # 将时间格式化成字符串
        convert(formatted_time)
    elif not date_pattern.match(formatted_time):
        print("请输入正确格式日期")
    else:
        convert(formatted_time)

if __name__ == '__main__':
    main()
